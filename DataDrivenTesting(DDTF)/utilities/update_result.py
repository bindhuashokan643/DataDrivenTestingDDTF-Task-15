import openpyxl
from datetime import datetime

def update_result(path, testid, result):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == testid:
            sheet.cell(row, 7).value = result
            sheet.cell(row, 4).value = datetime.now().strftime("%d-%m-%Y")
            sheet.cell(row, 5).value = datetime.now().strftime("%H:%M:%S")
    wb.save(path)
    print(f"Updated {testid} with result: {result}")
