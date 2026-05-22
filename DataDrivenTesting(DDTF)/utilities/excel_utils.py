import openpyxl
from datetime import datetime

def get_test_data(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append((row[0], row[1], row[2], row[5]))
    return data

def update_result(path, testid, result):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == testid:
            sheet.cell(row, 7).value = result
            sheet.cell(row, 4).value = datetime.now().strftime("%d-%m-%Y")
            sheet.cell(row, 5).value = datetime.now().strftime("%H:%M:%S")
    wb.save(path)
